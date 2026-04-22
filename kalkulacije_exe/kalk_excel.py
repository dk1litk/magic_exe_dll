"""kalk_excel.py — pretvori popis del CSV v formatiran XLSX (sheet BLIST).

Vhod: CSV (;-separated, windows-1250 / utf-8) s strukturo popisa del:
  R0: header (Z.st.;Sifra;Opis;Opomba;EM;Kolicina;Cena;Znesek)
  R1: naslov projekta (samo stolpec C)
  R2+: sekcije (3-stolpcni: A=hierarhija, B=st., C=naziv)
       ali postavke (6-stolpcni: A=hierarhija, B=st., C=opis, D=opomba, E=EM, F=kolicina)

Izhod XLSX:
  Sheet "BLIST", freeze A3, summary row zgoraj.
  Stolpci (7 vidnih + hidden Z):
    A = Nivo, B = Postavka, C = Opis postavke (opis + opombe),
    D = Enota mere, E = Količina, F = Cena, G = znesek,
    H..Y = rezervirani (prazni), Z = HIDDEN (alfa polje 10-15 znakov).
  R1: "Ponudbene postavke"
  R2: header
  R3: grand total (svetlo zelen BEF0BE, C=naslov iz CSV, G=SUM L1)
  Sekcije: L1 modra C6D7F5, L2..L7 zelena lestvica od temno do svetlo.
    Vse sekcije bold, G=SUM podrejenih.
  Postavke: brez fill, G=ROUND(E*F, 2); F (cena) prazno za uporabnika.

Uporaba: kalk_excel.exe <input.csv> <output.xlsx>
Exit codes: 0=OK, 1=napaka (brez outputa).
"""
import csv
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ENCODINGS = ("utf-8-sig", "utf-8", "cp1250", "cp1252")

# Fill za grand total vrstico (R3) — temno modra (opazno temnejsa od L1 C6D7F5).
FILL_TOTAL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
# L1 sekcija — modra.
COLOR_L1 = "FFC6D7F5"
# L2 (najtemnejsa zelena) -> L7 (svetlo zelena) interpolacija.
COLOR_L2_RGB = (108, 183, 108)  # #6CB76C
COLOR_L7_RGB = (232, 245, 232)  # #E8F5E8

FMT_KOLICINA = "#,##0.00##"
FMT_EUR = '#,##0.00\\ "€"'

# 8 vidnih stolpcev + Z (alfa polje 10-15 znakov, ne skrivamo po defaultu).
COL_WIDTHS = {
    "A": 22.0,   # Nivo (npr. "1.1.1.1.1.1.")
    "B": 10.0,   # Postavka (npr. "0001")
    "C": 50.0,   # Opis (pure, brez opomb)
    "D": 40.0,   # Opomba
    "E": 11.0,   # Enota mere (m2, KOS, M3, ...)
    "F": 10.0,   # Kolicina
    "G": 12.0,   # Cena
    "H": 14.0,   # Znesek
    "Z": 15.0,   # alfa polje (koda zunanjega sistema)
}

HEADERS = ["Nivo", "Postavka", "Opis", "Opomba", "Enota mere", "Količina", "Cena", "znesek"]

# Stolpec formul: H = znesek. Za ROUND pri postavkah: H = ROUND(F * G, 2).
COL_ZNESEK = "H"
COL_KOLICINA = "F"
COL_CENA = "G"

# Stevilski-format stolpci (1-based).
COL_IDX_KOLICINA = 6  # F
COL_IDX_CENA = 7      # G
COL_IDX_ZNESEK = 8    # H
NUM_VISIBLE_COLS = 8  # A..H

WRAP = Alignment(wrap_text=True, vertical="top")


def fill_for_level(level: int) -> PatternFill:
    """Barva ozadja za sekcijo danega nivoja.

    L1 = modra. L2..L7 = zelena lestvica od temno (L2) do svetlo (L7).
    Vec kot 7 sekcijskih nivojev se ne pricakuje — ce se zgodi, dobi barvo L7.
    """
    if level <= 1:
        return PatternFill(start_color=COLOR_L1, end_color=COLOR_L1, fill_type="solid")
    # L2..L7 -> t = 0.0 .. 1.0
    clamped = min(7, level)
    t = (clamped - 2) / 5.0  # L2=0, L7=1
    r2, g2, b2 = COLOR_L2_RGB
    r7, g7, b7 = COLOR_L7_RGB
    r = round(r2 + t * (r7 - r2))
    g = round(g2 + t * (g7 - g2))
    b = round(b2 + t * (b7 - b2))
    color = f"FF{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


@dataclass
class Row:
    kind: str  # "section" | "item"
    level: int  # level sekcije (stevilo pik v A); pri postavki = leaf_section.level + 1
    a: str
    b: str
    c: str
    d: str = ""
    e: str = ""
    f: float | None = None
    z: str = ""  # 26. stolpec CSV -> stolpec Z v XLSX (alfa koda za zunanji sistem)
    xlsx_row: int = 0
    children_rows: list[int] = field(default_factory=list)


def decode(raw: bytes) -> str:
    for enc in ENCODINGS:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("ne morem dekodirati CSV")


def parse_num(s: str) -> float | None:
    s = s.strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def parse_csv(path: Path) -> tuple[str, list[Row]]:
    raw = path.read_bytes()
    text = decode(raw)
    reader = list(csv.reader(text.splitlines(), delimiter=";", quotechar='"'))

    if len(reader) < 2:
        raise ValueError("CSV ima premalo vrstic")

    title = reader[1][2].strip() if len(reader[1]) >= 3 else ""

    rows: list[Row] = []
    for raw_row in reader[2:]:
        if not raw_row or not raw_row[0].strip():
            continue
        a = raw_row[0].strip()
        level = a.count(".")  # "1.1.1." = 3 pike -> L3
        if len(raw_row) == 3:
            rows.append(Row(kind="section", level=level, a=a,
                            b=raw_row[1].strip(), c=raw_row[2].strip()))
        elif len(raw_row) >= 5:
            b = raw_row[1].strip()
            c = raw_row[2].strip()
            d = raw_row[3].strip() if len(raw_row) > 3 else ""
            e = raw_row[4].strip() if len(raw_row) > 4 else ""
            f = parse_num(raw_row[5]) if len(raw_row) > 5 else None
            # 26. stolpec (indeks 25) = alfa koda za zunanji sistem (stolpec Z v XLSX).
            z = raw_row[25].strip() if len(raw_row) > 25 else ""
            rows.append(Row(kind="item", level=level + 1, a=a,
                            b=b, c=c, d=d, e=e, f=f, z=z))

    return title, rows


def assign_xlsx_rows_and_children(rows: list[Row]) -> None:
    """xlsx_row zacne pri 4. Za vsako sekcijo sestavi children_rows (direct children)."""
    for i, r in enumerate(rows):
        r.xlsx_row = 4 + i

    for i, r in enumerate(rows):
        if r.kind != "section":
            continue
        children: list[int] = []
        for j in range(i + 1, len(rows)):
            nxt = rows[j]
            if nxt.kind == "section" and nxt.level <= r.level:
                break
            # Direct children (subsekcije ali postavke na nivoju r.level + 1).
            if nxt.level == r.level + 1:
                children.append(nxt.xlsx_row)
        r.children_rows = children


def sum_formula(rows_xlsx: list[int]) -> str:
    if not rows_xlsx:
        return ""
    rows_xlsx = sorted(rows_xlsx)
    if len(rows_xlsx) > 1 and all(rows_xlsx[i] + 1 == rows_xlsx[i + 1]
                                   for i in range(len(rows_xlsx) - 1)):
        return f"=SUM({COL_ZNESEK}{rows_xlsx[0]}:{COL_ZNESEK}{rows_xlsx[-1]})"
    refs = ",".join(f"{COL_ZNESEK}{r}" for r in rows_xlsx)
    return f"=SUM({refs},)"


_OPOMBA_RE = re.compile(r"(?<!^)(?<!\n)(Opomba:|Opombe:)")
_OPOMBA_START_RE = re.compile(r"(Opomba:|Opombe:)")


def _split_opombe(text: str) -> str:
    """Vstavi newline pred vsakim 'Opomba:' / 'Opombe:' markerjem znotraj teksta."""
    if not text:
        return text
    return _OPOMBA_RE.sub(r"\n\1", text)


def split_opis_opomba(r: Row) -> tuple[str, str]:
    """Loci Opis (stolpec C) in Opomba (stolpec D).

    1. Ce CSV Opomba (r.d) ni prazna: C = r.c, D = r.d (oboje s splitom markerjev).
    2. Sicer poglej v r.c za 'Opomba:' / 'Opombe:' — vse pred markerjem gre v C,
       ostalo z markerjem vred v D.
    3. Ce markerja ni: vse v C, D prazen.
    """
    c_raw = r.c or ""
    d_raw = r.d or ""

    if d_raw.strip():
        return _split_opombe(c_raw), _split_opombe(d_raw)

    m = _OPOMBA_START_RE.search(c_raw)
    if m:
        opis = c_raw[:m.start()].rstrip(" .,;:-")
        opomba = _split_opombe(c_raw[m.start():])
        return opis, opomba

    return c_raw, ""


def l1_xlsx_rows(rows: list[Row]) -> list[int]:
    return [r.xlsx_row for r in rows if r.kind == "section" and r.level == 1]


def apply_cell_defaults(ws, row: int, fill: PatternFill | None, make_bold: bool) -> None:
    """Aplicira alignment/bold/fill in number formate na vsa vidna polja vrstice."""
    bold_font = Font(bold=True)
    for c_idx in range(1, NUM_VISIBLE_COLS + 1):  # A..H
        cell = ws.cell(row=row, column=c_idx)
        cell.alignment = WRAP
        if make_bold:
            cell.font = bold_font
        if fill is not None:
            cell.fill = fill
    ws.cell(row=row, column=COL_IDX_KOLICINA).number_format = FMT_KOLICINA  # F
    ws.cell(row=row, column=COL_IDX_CENA).number_format = FMT_EUR            # G
    ws.cell(row=row, column=COL_IDX_ZNESEK).number_format = FMT_EUR          # H


def write_xlsx(title: str, rows: list[Row], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BLIST"

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # Column widths (A..G + Z).
    for letter, w in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    # R1: glavni naslov
    ws.cell(row=1, column=1, value="Ponudbene postavke").alignment = WRAP

    # R2: header
    for c_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.alignment = WRAP
    ws.cell(row=2, column=COL_IDX_KOLICINA).number_format = FMT_KOLICINA
    ws.cell(row=2, column=COL_IDX_CENA).number_format = FMT_EUR
    ws.cell(row=2, column=COL_IDX_ZNESEK).number_format = FMT_EUR

    # R3: grand total
    apply_cell_defaults(ws, 3, FILL_TOTAL, make_bold=True)
    ws.cell(row=3, column=3, value=title)
    total_children = l1_xlsx_rows(rows)
    if total_children:
        ws.cell(row=3, column=COL_IDX_ZNESEK, value=sum_formula(total_children))

    # R4+: sekcije in postavke
    for r in rows:
        xr = r.xlsx_row

        ws.cell(row=xr, column=1, value=r.a)  # Nivo
        ws.cell(row=xr, column=2, value=r.b)  # Postavka

        if r.kind == "item":
            opis, opomba = split_opis_opomba(r)
            ws.cell(row=xr, column=3, value=opis)
            ws.cell(row=xr, column=4, value=opomba)
            ws.cell(row=xr, column=5, value=r.e)  # Enota mere
            if r.f is not None:
                ws.cell(row=xr, column=6, value=r.f)  # Kolicina
            # G (Cena) prazno — izpolni uporabnik.
            ws.cell(row=xr, column=COL_IDX_ZNESEK,
                    value=f"=ROUND(${COL_KOLICINA}{xr}*{COL_CENA}{xr},2)")
            if r.z:
                ws.cell(row=xr, column=26, value=r.z)
            fill = None
            make_bold = False
        else:  # section
            ws.cell(row=xr, column=3, value=r.c)
            formula = sum_formula(r.children_rows)
            if formula:
                ws.cell(row=xr, column=COL_IDX_ZNESEK, value=formula)
            fill = fill_for_level(r.level)
            make_bold = True

        apply_cell_defaults(ws, xr, fill, make_bold)

        # Row outline level (gumbi levo zgoraj).
        ws.row_dimensions[xr].outline_level = r.level

    ws.freeze_panes = "A3"
    wb.save(out_path)


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        return 1

    in_path = Path(argv[1])
    out_path = Path(argv[2])

    if not in_path.is_file():
        return 1

    try:
        title, rows = parse_csv(in_path)
    except Exception:
        return 1

    assign_xlsx_rows_and_children(rows)

    try:
        write_xlsx(title, rows, out_path)
    except Exception:
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
